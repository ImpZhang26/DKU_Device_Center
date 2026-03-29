"""
设备选配应用 - 视图函数
Django 4.2 views
"""
import json
import hashlib
import random
import os
from datetime import datetime
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import models
from django.db.models import Count, Sum
from django.utils import timezone
from django import template

from .models import (
    ProductType, ProductModel, ProductConfiguration,
    Accessory, CartItem, Order, Admin, DellProduct, LenovoProduct,
    ProductImage
)
from .email_utils import send_order_email


# 注册模板标签
register = template.Library()

@register.filter
def split(value, arg):
    """分割字符串"""
    if value:
        return value.split(arg)
    return []


def get_user_key(request):
    """获取用户标识（优先使用用户ID，否则使用session）"""
    if request.user.is_authenticated:
        return request.user.username
    if 'user_key' not in request.session:
        request.session['user_key'] = f"guest_{random.randint(100000, 999999)}"
    return request.session['user_key']


def global_settings(request):
    """全局上下文处理器"""
    return {
        'brand': 'Device Center',
        'current_year': datetime.now().year,
    }


# ==================== 首页和品牌选择 ====================

def index(request):
    """首页 - 品牌选择"""
    return render(request, 'index.html')


# ==================== 认证相关 ====================

def login_view(request):
    """登录页面"""
    if request.method == 'POST':
        login_type = request.POST.get('login_type', 'netid')
        
        if login_type == 'admin':
            # 管理员登录
            username = request.POST.get('username', '')
            password = request.POST.get('password', '')
            
            try:
                admin = Admin.objects.get(username=username, is_active=True)
                password_hash = hashlib.sha256(password.encode()).hexdigest()
                
                if admin.password_hash == password_hash:
                    # 创建session
                    request.session['admin_id'] = admin.id
                    request.session['admin_name'] = admin.name
                    request.session['is_admin'] = True
                    return redirect('manage')
                else:
                    return render(request, 'login.html', {'error': '密码错误'})
            except Admin.DoesNotExist:
                return render(request, 'login.html', {'error': '管理员账号不存在'})
        
        else:
            # NETID登录
            netid = request.POST.get('netid', '').strip()
            
            if netid:
                # 创建或更新session
                request.session['user_id'] = netid
                request.session['user_name'] = netid
                request.session['is_admin'] = False
                request.session.modified = True
                return redirect('apple_index')
            
            return render(request, 'login.html', {'error': '请输入NETID'})
    
    return render(request, 'login.html')


def logout_view(request):
    """退出登录"""
    logout(request)
    return redirect('login')


# ==================== Apple 选配中心 ====================

def apple_index(request):
    """Apple选配中心首页"""
    product_types = ProductType.objects.filter(is_active=True)
    return render(request, 'apple/index.html', {'product_types': product_types})


def apple_laptop(request):
    """MacBook选配"""
    device_type = ProductType.objects.get(device_type='Laptop')
    models = ProductModel.objects.filter(device_type=device_type, is_active=True)
    
    # 获取选中的型号ID
    selected_model_id = request.GET.get('model')
    selected_model = None
    configurations = []
    
    if selected_model_id:
        try:
            selected_model = ProductModel.objects.get(id=selected_model_id)
            configurations = ProductConfiguration.objects.filter(
                model=selected_model, is_active=True
            ).order_by('discount_price')
        except ProductModel.DoesNotExist:
            pass
    
    return render(request, 'apple/laptop.html', {
        'models': models,
        'selected_model': selected_model,
        'configurations': configurations,
        'brand': 'Apple',
    })


def apple_desktop(request):
    """台式机选配"""
    device_type = ProductType.objects.get(device_type='Desktop')
    models = ProductModel.objects.filter(device_type=device_type, is_active=True)
    
    selected_model_id = request.GET.get('model')
    selected_model = None
    configurations = []
    
    if selected_model_id:
        try:
            selected_model = ProductModel.objects.get(id=selected_model_id)
            configurations = ProductConfiguration.objects.filter(
                model=selected_model, is_active=True
            ).order_by('discount_price')
        except ProductModel.DoesNotExist:
            pass
    
    return render(request, 'apple/desktop.html', {
        'models': models,
        'selected_model': selected_model,
        'configurations': configurations,
        'brand': 'Apple',
    })


def apple_ipad(request):
    """iPad选配"""
    device_type = ProductType.objects.get(device_type='iPad')
    models = ProductModel.objects.filter(device_type=device_type, is_active=True)
    
    selected_model_id = request.GET.get('model')
    selected_model = None
    configurations = []
    
    if selected_model_id:
        try:
            selected_model = ProductModel.objects.get(id=selected_model_id)
            configurations = ProductConfiguration.objects.filter(
                model=selected_model, is_active=True
            ).order_by('discount_price')
        except ProductModel.DoesNotExist:
            pass
    
    return render(request, 'apple/ipad.html', {
        'models': models,
        'selected_model': selected_model,
        'configurations': configurations,
        'brand': 'Apple',
    })


def apple_accessories(request):
    """配件页面"""
    accessories = Accessory.objects.filter(is_active=True)
    
    # 按类型分组
    accessory_groups = {}
    for acc in accessories:
        if acc.device_type not in accessory_groups:
            accessory_groups[acc.device_type] = []
        accessory_groups[acc.device_type].append(acc)
    
    return render(request, 'apple/accessories.html', {
        'accessory_groups': accessory_groups,
        'brand': 'Apple',
    })


def apple_detail(request, config_id):
    """产品详情页"""
    try:
        config = ProductConfiguration.objects.select_related('model__device_type').get(id=config_id)
    except ProductConfiguration.DoesNotExist:
        return render(request, '404.html', status=404)
    
    device_type = config.model.device_type.device_type
    device_type_name = config.model.device_type.name
    
    # 计算折扣百分比
    if config.original_price and config.original_price > config.discount_price:
        discount_percent = int((1 - config.discount_price / config.original_price) * 100)
    else:
        discount_percent = 0
    
    # 获取产品图片 - 从 ProductImage 模型查询
    main_image = None
    thumbnails = []
    
    # 查询该配置的图片
    images = ProductImage.objects.filter(
        brand='apple',
        product_type='config',
        product_id=config_id
    ).order_by('sort_order', 'created_at')
    
    # 如果没有配置级图片，尝试查询型号级图片
    if not images.exists():
        images = ProductImage.objects.filter(
            brand='apple',
            product_type='model',
            product_id=config.model.id
        ).order_by('sort_order', 'created_at')
    
    if images.exists():
        # 获取主图（优先显示主图，然后是第一张图）
        main_img_obj = images.filter(image_type='main').first() or images.first()
        if main_img_obj and main_img_obj.image:
            main_image = main_img_obj.image.url
        
        # 获取所有图片作为缩略图
        for img in images:
            if img.image:
                thumbnails.append(img.image.url)
    else:
        # 如果没有上传图片，使用型号的图片作为后备
        if config.model.image_url:
            main_image = config.model.image_url
    
    # 处理configuration_parameters - 转换为HTML格式
    configuration_parameters_html = ''
    if config.configuration_parameters:
        params = config.configuration_parameters.replace('；', ';').split(';')
        configuration_parameters_html = '<ul>'
        for param in params:
            param = param.strip()
            if param:
                configuration_parameters_html += f'<li>{param}</li>'
        configuration_parameters_html += '</ul>'
    
    # 处理changing_config_price - 转换为HTML格式
    configuration_upgrade_html = ''
    if config.changing_config_price:
        lines = config.changing_config_price.split('\n')
        configuration_upgrade_html = '<div class="upgrade-list">'
        
        for line in lines:
            line = line.strip()
            if line:
                # 处理特殊符号开头的内容
                upgrade_name = line
                if line.startswith('★'):
                    upgrade_name = line[1:].strip()
                
                # 尝试提取价格（格式如 "+1200 yuan" 或 "，+1200 yuan"）
                price_parts = upgrade_name.rsplit('，+', 1)
                if len(price_parts) == 2:
                    name = price_parts[0].strip()
                    price = '+' + price_parts[1].strip()
                    configuration_upgrade_html += f'''<div class="upgrade-item">
                        <div class="upgrade-checkbox"></div>
                        <div class="upgrade-info">
                            <span class="upgrade-name">{name}</span>
                            <span class="upgrade-price">{price}</span>
                        </div>
                    </div>'''
                elif '+' in upgrade_name:
                    # 直接在末尾的格式
                    parts = upgrade_name.rsplit('+', 1)
                    if len(parts) == 2:
                        name = parts[0].strip()
                        price = '+' + parts[1].strip()
                        configuration_upgrade_html += f'''<div class="upgrade-item">
                            <div class="upgrade-checkbox"></div>
                            <div class="upgrade-info">
                                <span class="upgrade-name">{name}</span>
                                <span class="upgrade-price">{price}</span>
                            </div>
                        </div>'''
                    else:
                        configuration_upgrade_html += f'<div class="upgrade-item">{upgrade_name}</div>'
                else:
                    configuration_upgrade_html += f'<div class="upgrade-item">{upgrade_name}</div>'
        
        configuration_upgrade_html += '</div>'
    
    # 设置用户信息
    user_name = request.session.get('user_name', '')
    
    return render(request, 'apple/detail.html', {
        'configuration': config,
        'device_type': device_type,
        'device_type_name': device_type_name,
        'brand': 'Apple',
        'user_name': user_name,
        'discount_percent': discount_percent,
        'configuration_parameters_html': configuration_parameters_html,
        'configuration_upgrade_html': configuration_upgrade_html,
        'main_image': main_image,
        'thumbnails': thumbnails,
    })


# ==================== API 接口 ====================

@csrf_exempt
def api_product_types(request, brand):
    """获取产品类型列表"""
    types = ProductType.objects.filter(is_active=True)
    data = [{'device_type': t.device_type, 'name': t.name, 'description': t.description} 
            for t in types]
    return JsonResponse({'success': True, 'data': data})


@csrf_exempt
def api_models(request, brand, device_type):
    """获取型号列表"""
    try:
        ptype = ProductType.objects.get(device_type=device_type)
        models = ProductModel.objects.filter(device_type=ptype, is_active=True)
        data = [{'id': m.id, 'model_name': m.model_name, 'description': m.description} 
                for m in models]
        return JsonResponse({'success': True, 'data': data})
    except ProductType.DoesNotExist:
        return JsonResponse({'success': False, 'message': '产品类型不存在'})


@csrf_exempt
def api_configurations(request, brand, model_id):
    """获取配置列表"""
    try:
        configs = ProductConfiguration.objects.filter(
            model_id=model_id, is_active=True
        ).order_by('discount_price')
        data = [{
            'id': c.id,
            'configuration': c.configuration,
            'cpu': c.cpu,
            'memory': c.memory,
            'hard_disk': c.hard_disk,
            'graphic_card': c.graphic_card,
            'screen': c.screen,
            'original_price': str(c.original_price),
            'discount_price': str(c.discount_price),
            'discount_rate': str(c.discount_rate),
        } for c in configs]
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
def api_accessories(request, brand):
    """获取配件列表"""
    accessories = Accessory.objects.filter(is_active=True)
    data = [{
        'id': a.id,
        'model': a.model,
        'device_type': a.device_type,
        'description': a.description,
        'original_price': str(a.original_price),
        'discount_price': str(a.discount_price),
        'discount_rate': str(a.discount_rate),
    } for a in accessories]
    return JsonResponse({'success': True, 'data': data})


# ==================== 购物车 ====================

@csrf_exempt
@require_http_methods(["POST"])
def api_cart_add(request, brand):
    """添加购物车"""
    try:
        data = json.loads(request.body)
        user_key = get_user_key(request)
        
        cart_item = CartItem.objects.create(
            user_id=user_key,
            brand=brand,
            device_type=data.get('device_type', ''),
            model_name=data.get('model_name', ''),
            configuration_id=data.get('configuration_id'),
            selected_config=data,
        )
        
        return JsonResponse({'success': True, 'message': '添加成功', 'cart_id': cart_item.id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def api_cart_remove(request, brand, cart_id):
    """移除购物车项"""
    try:
        CartItem.objects.filter(id=cart_id, user_id=get_user_key(request)).delete()
        return JsonResponse({'success': True, 'message': '已移除'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
def api_cart_list(request, brand):
    """获取购物车列表"""
    user_key = get_user_key(request)
    items = CartItem.objects.filter(user_id=user_key, brand=brand)
    data = [{
        'id': item.id,
        'model_name': item.model_name,
        'selected_config': item.selected_config,
        'quantity': item.quantity,
        'created_at': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
    } for item in items]
    return JsonResponse({'success': True, 'data': data})


# ==================== 订单 ====================

@csrf_exempt
@require_http_methods(["POST"])
def api_order_create(request, brand):
    """创建订单"""
    try:
        data = json.loads(request.body)
        
        # 检查用户是否已通过NETID登录
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({
                'success': False, 
                'message': '请先登录后再提交订单',
                'code': 'not_logged_in'
            }, status=401)
        
        user_key = user_id  # 使用已登录的用户ID
        
        # 生成订单号
        order_number = f"{brand.upper()}{timezone.now().strftime('%Y%m%d%H%M%S')}{random.randint(1000, 9999)}"
        
        # 处理升级选项
        selected_upgrades = data.get('selected_upgrades', [])
        upgrade_price = data.get('upgrade_price', 0)
        base_price = data.get('base_price', 0)
        
        # 将升级选项转换为JSON字符串存储
        upgrades_json = json.dumps(selected_upgrades, ensure_ascii=False) if selected_upgrades else ''
        
        order = Order.objects.create(
            order_number=order_number,
            user_id=user_key,
            user_name=request.session.get('user_name', ''),
            user_email=request.session.get('user_email', ''),
            brand=brand,
            device_type=data.get('device_type', ''),
            model_name=data.get('model_name', ''),
            selected_cpu=data.get('cpu', ''),
            selected_memory=data.get('memory', ''),
            selected_hard_disk=data.get('hard_disk', ''),
            selected_graphic_card=data.get('graphic_card', ''),
            selected_screen=data.get('screen', ''),
            base_price=base_price,
            upgrade_price=upgrade_price,
            selected_upgrades=upgrades_json,
            total_price=data.get('total_price', 0),
            remark=data.get('remark', ''),
        )

        # 发送订单确认邮件
        order_email_data = {
            'order_number': order_number,
            'user_id': user_key,
            'user_name': request.session.get('user_name', ''),
            'user_email': request.session.get('user_email', ''),
            'brand': brand,
            'model_name': data.get('model_name', ''),
            'cpu': data.get('cpu', ''),
            'memory': data.get('memory', ''),
            'hard_disk': data.get('hard_disk', ''),
            'graphic_card': data.get('graphic_card', ''),
            'screen': data.get('screen', ''),
            'base_price': base_price,
            'upgrade_price': upgrade_price,
            'total_price': data.get('total_price', 0),
            'selected_upgrades': selected_upgrades,
            'status': order.status,
            'created_at': order.created_at,
        }
        send_order_email(order_email_data)

        return JsonResponse({'success': True, 'message': '订单创建成功', 'order_number': order_number})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
def api_orders(request, brand):
    """获取订单列表"""
    user_key = get_user_key(request)
    orders = Order.objects.filter(user_id=user_key, brand=brand)
    data = [{
        'id': o.id,
        'order_number': o.order_number,
        'model_name': o.model_name,
        'total_price': str(o.total_price),
        'status': o.status,
        'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S'),
    } for o in orders]
    return JsonResponse({'success': True, 'data': data})


# ==================== 我的订单 ====================

@require_http_methods(["GET"])
def my_orders(request):
    """我的订单页面 - 包含所有品牌的订单"""
    # 检查用户是否已通过NETID登录
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')
    
    user_name = request.session.get('user_name', '')
    
    # 获取用户的所有订单（不区分品牌）
    orders = Order.objects.filter(user_id=user_id).order_by('-created_at')
    
    # 统计各状态订单数量
    all_count = orders.count()
    pending_count = orders.filter(status='pending').count()
    processing_count = orders.filter(status='processing').count()
    completed_count = orders.filter(status='completed').count()
    cancelled_count = orders.filter(status='cancelled').count()
    
    return render(request, 'orders.html', {
        'orders': orders,
        'user_name': user_name,
        'all_count': all_count,
        'pending_count': pending_count,
        'processing_count': processing_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
    })


@csrf_exempt
@require_http_methods(["POST"])
def api_order_cancel(request):
    """取消/终止订单"""
    try:
        data = json.loads(request.body)
        order_number = data.get('order_number', '')
        user_key = get_user_key(request)
        
        # 验证订单属于当前用户
        order = Order.objects.get(order_number=order_number, user_id=user_key)
        
        # 只有待处理或处理中的订单可以取消
        if order.status not in ['pending', 'processing']:
            return JsonResponse({
                'success': False, 
                'message': '当前订单状态不允许取消'
            })
        
        # 更新订单状态为已取消
        order.status = 'cancelled'
        order.remark = (order.remark + ' | 用户主动终止订单' if order.remark else '用户主动终止订单')
        order.save()
        
        return JsonResponse({'success': True, 'message': '订单已成功终止'})
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'message': '订单不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== 管理后台 ====================

def require_admin(view_func):
    """管理员权限装饰器 - 支持API和浏览器请求"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # 检查是否已登录为管理员
        if not request.session.get('is_admin'):
            # 判断是否为API/AJAX请求
            # 1. 检查X-Requested-With头
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            # 2. 检查请求路径是否以/api/开头
            is_api_path = request.path.startswith('/api/')
            # 3. 检查Accept头是否包含application/json
            accept_json = 'application/json' in request.headers.get('Accept', '')
            
            if is_ajax or is_api_path or accept_json:
                # 返回JSON 401响应
                return JsonResponse({
                    'success': False, 
                    'message': '请先登录管理员账号',
                    'code': 'unauthorized'
                }, status=401)
            
            # 浏览器请求，重定向到登录页
            return redirect('login')
        
        # 已登录，放行请求
        return view_func(request, *args, **kwargs)
    return wrapper


@require_admin
def admin_dashboard(request):
    """管理后台首页"""
    # 统计信息
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    total_users = Order.objects.values('user_id').distinct().count()
    total_revenue = Order.objects.filter(status='completed').aggregate(
        total=models.Sum('total_price')
    )['total'] or 0
    
    # 近期订单
    recent_orders = Order.objects.all()[:10]
    
    # 按品牌统计
    brand_stats = Order.objects.values('brand').annotate(
        count=Count('id'),
        total=models.Sum('total_price')
    )
    
    return render(request, 'admin/dashboard.html', {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'total_users': total_users,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
        'brand_stats': brand_stats,
    })


@csrf_exempt
@require_admin
def api_admin_orders(request):
    """管理后台 - 订单列表"""
    orders = Order.objects.all().order_by('-created_at')
    data = [{
        'id': o.id,
        'order_number': o.order_number,
        'user_id': o.user_id,
        'user_name': o.user_name,
        'user_email': o.user_email,
        'brand': o.brand,
        'device_type': o.device_type,
        'model_name': o.model_name,
        'total_price': str(o.total_price),
        'status': o.status,
        'remark': o.remark,
        'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S'),
    } for o in orders]
    return JsonResponse({'success': True, 'data': data})


@csrf_exempt
@require_admin
def api_admin_order_detail(request, order_id):
    """管理后台 - 订单详情"""
    try:
        order = Order.objects.get(id=order_id)
        
        # 解析升级选项JSON
        selected_upgrades = []
        if order.selected_upgrades:
            if isinstance(order.selected_upgrades, str):
                import json
                selected_upgrades = json.loads(order.selected_upgrades)
            else:
                selected_upgrades = order.selected_upgrades
        
        data = {
            'id': order.id,
            'order_number': order.order_number,
            'user_id': order.user_id,
            'user_name': order.user_name,
            'user_email': order.user_email,
            'brand': order.brand,
            'device_type': order.device_type,
            'model_name': order.model_name,
            'selected_cpu': order.selected_cpu,
            'selected_memory': order.selected_memory,
            'selected_hard_disk': order.selected_hard_disk,
            'selected_graphic_card': order.selected_graphic_card,
            'selected_screen': order.selected_screen,
            'base_price': str(order.base_price),
            'upgrade_price': str(order.upgrade_price),
            'selected_upgrades': selected_upgrades,
            'total_price': str(order.total_price),
            'status': order.status,
            'remark': order.remark,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': order.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
        }
        return JsonResponse({'success': True, 'data': data})
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'message': '订单不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_order_status(request, order_id):
    """管理后台 - 更新订单状态"""
    try:
        order = Order.objects.get(id=order_id)
        data = json.loads(request.body)
        order.status = data.get('status', order.status)
        order.save()
        return JsonResponse({'success': True, 'message': '状态更新成功'})
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'message': '订单不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_stats(request):
    """管理后台 - 统计数据"""
    # 今日订单
    today = timezone.now().date()
    today_orders = Order.objects.filter(created_at__date=today)

    # 订单状态分布
    status_dist = Order.objects.values('status').annotate(count=Count('id'))

    return JsonResponse({
        'success': True,
        'data': {
            'today_orders': today_orders.count(),
            'today_revenue': today_orders.aggregate(total=models.Sum('total_price'))['total'] or 0,
            'status_distribution': {s['status']: s['count'] for s in status_dist},
        }
    })


@csrf_exempt
@require_admin
def export_orders_excel(request):
    """管理后台 - 导出订单到Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # 获取所有订单，按创建时间倒序
        orders = Order.objects.all().order_by('-created_at')
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "订单数据"
        
        # 定义表头
        headers = [
            '订单号',
            '用户ID',
            '用户姓名',
            '用户邮箱',
            '品牌',
            '产品型号',
            'CPU',
            '内存',
            '硬盘',
            '显卡',
            '屏幕',
            '基础价格(元)',
            '升级费用(元)',
            '升级选项',
            '总价(元)',
            '状态',
            '创建时间',
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 添加表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # 设置列宽
            ws.column_dimensions[chr(64 + col)].width = 18
        
        # 状态映射（中文字符串）
        status_map = {
            'pending': '待处理',
            'processing': '处理中',
            'completed': '已完成',
            'cancelled': '已终止',
        }
        
        # 添加数据行
        for row_idx, order in enumerate(orders, 2):
            # 解析升级选项
            selected_upgrades_text = ''
            if order.selected_upgrades:
                try:
                    if isinstance(order.selected_upgrades, str):
                        upgrades_list = json.loads(order.selected_upgrades)
                    else:
                        upgrades_list = order.selected_upgrades
                    if upgrades_list:
                        upgrade_names = [u.get('name', str(u)) for u in upgrades_list]
                        selected_upgrades_text = ', '.join(upgrade_names)
                except (json.JSONDecodeError, TypeError):
                    selected_upgrades_text = str(order.selected_upgrades)
            
            row_data = [
                order.order_number,
                order.user_id,
                order.user_name,
                order.user_email,
                order.brand.upper() if order.brand else '',
                order.model_name,
                order.selected_cpu or '',
                order.selected_memory or '',
                order.selected_hard_disk or '',
                order.selected_graphic_card or '',
                order.selected_screen or '',
                str(order.base_price) if order.base_price else '0',
                str(order.upgrade_price) if order.upgrade_price else '0',
                selected_upgrades_text,
                str(order.total_price) if order.total_price else '0',
                status_map.get(order.status, order.status),
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为所有数据单元格添加边框
        for row in ws.iter_rows(min_row=1, max_row=len(orders) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # 生成文件名
        filename = f"订单数据_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 创建HTTP响应，设置正确的Content-Type
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 保存工作簿到响应
        wb.save(response)
        
        return response
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': '请安装 openpyxl 库: pip install openpyxl'
        }, status=500)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'导出失败: {str(e)}'
        }, status=500)


# ============ 产品管理 API ============

# Apple 产品型号管理
@csrf_exempt
@require_admin
def api_admin_apple_models(request):
    """获取/添加 Apple 产品型号列表"""
    if request.method == 'GET':
        models = ProductModel.objects.select_related('device_type').all().order_by('device_type', 'model_name')
        data = [{
            'id': m.id,
            'model_name': m.model_name,
            'device_type': m.device_type.device_type,
            'device_type_name': m.device_type.name,
            'description': m.description,
            'image_url': m.image_url,
            'is_active': m.is_active,
            'config_count': m.configurations.count(),
        } for m in models]
        return JsonResponse({'success': True, 'data': data})
    elif request.method == 'POST':
        data = json.loads(request.body)
        try:
            device_type = ProductType.objects.get(device_type=data.get('device_type'))
            model = ProductModel.objects.create(
                model_name=data.get('model_name'),
                device_type=device_type,
                description=data.get('description', ''),
                image_url=data.get('image_url', ''),
            )
            return JsonResponse({'success': True, 'message': '型号创建成功', 'id': model.id})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_admin
def api_admin_apple_model(request, model_id):
    """编辑/删除 Apple 产品型号"""
    try:
        model = ProductModel.objects.get(id=model_id)
    except ProductModel.DoesNotExist:
        return JsonResponse({'success': False, 'message': '型号不存在'})

    if request.method == 'POST':
        data = json.loads(request.body)
        model.model_name = data.get('model_name', model.model_name)
        model.description = data.get('description', model.description)
        model.image_url = data.get('image_url', model.image_url)
        model.is_active = data.get('is_active', model.is_active)
        model.save()
        return JsonResponse({'success': True, 'message': '更新成功'})
    elif request.method == 'DELETE':
        model.delete()
        return JsonResponse({'success': True, 'message': '删除成功'})

@csrf_exempt
@require_admin
def api_admin_apple_configs(request):
    """添加 Apple 产品配置"""
    data = json.loads(request.body)
    try:
        model = ProductModel.objects.get(id=data.get('model_id'))
        config = ProductConfiguration.objects.create(
            model=model,
            configuration=data.get('configuration'),
            cpu=data.get('cpu', ''),
            memory=data.get('memory', ''),
            hard_disk=data.get('hard_disk', ''),
            graphic_card=data.get('graphic_card', ''),
            screen=data.get('screen', ''),
            original_price=data.get('original_price', 0),
            discount_price=data.get('discount_price', 0),
        )
        return JsonResponse({'success': True, 'message': '配置创建成功', 'id': config.id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_admin
def api_admin_apple_config(request, config_id):
    """编辑/删除 Apple 产品配置"""
    try:
        config = ProductConfiguration.objects.get(id=config_id)
    except ProductConfiguration.DoesNotExist:
        return JsonResponse({'success': False, 'message': '配置不存在'})

    if request.method == 'POST':
        data = json.loads(request.body)
        config.configuration = data.get('configuration', config.configuration)
        config.cpu = data.get('cpu', config.cpu)
        config.memory = data.get('memory', config.memory)
        config.hard_disk = data.get('hard_disk', config.hard_disk)
        config.graphic_card = data.get('graphic_card', config.graphic_card)
        config.screen = data.get('screen', config.screen)
        config.original_price = data.get('original_price', config.original_price)
        config.discount_price = data.get('discount_price', config.discount_price)
        config.is_active = data.get('is_active', config.is_active)
        config.save()
        return JsonResponse({'success': True, 'message': '更新成功'})
    elif request.method == 'DELETE':
        config.delete()
        return JsonResponse({'success': True, 'message': '删除成功'})

# Dell 产品管理
@csrf_exempt
@require_admin
def api_admin_dell_products(request):
    """获取/添加 Dell 产品列表"""
    if request.method == 'GET':
        products = DellProduct.objects.all().order_by('device_type', 'model')
        data = [{
            'id': p.id,
            'device_type': p.device_type,
            'model': p.model,
            'basic_config': p.basic_config,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'price': str(p.price),
            'original_price': str(p.original_price),
            'image_url': p.image_url,
            'is_active': p.is_active,
        } for p in products]
        return JsonResponse({'success': True, 'data': data})
    elif request.method == 'POST':
        data = json.loads(request.body)
        try:
            product = DellProduct.objects.create(
                device_type=data.get('device_type'),
                model=data.get('model'),
                basic_config=data.get('basic_config', ''),
                cpu=data.get('cpu', ''),
                memory=data.get('memory', ''),
                hard_disk=data.get('hard_disk', ''),
                graphic_card=data.get('graphic_card', ''),
                wifi=data.get('wifi', ''),
                configuration_parameters=data.get('configuration_parameters', ''),
                weight=data.get('weight', ''),
                price=data.get('price', 0),
                original_price=data.get('original_price', 0),
                discount_rate=data.get('discount_rate', 0),
                changing_config_price=data.get('changing_config_price', ''),
                official_link=data.get('official_link', ''),
                image_url=data.get('image_url', ''),
            )
            return JsonResponse({'success': True, 'message': '产品创建成功', 'id': product.id})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_admin
def api_admin_dell_product(request, product_id):
    """编辑/删除 Dell 产品"""
    try:
        product = DellProduct.objects.get(id=product_id)
    except DellProduct.DoesNotExist:
        return JsonResponse({'success': False, 'message': '产品不存在'})

    if request.method == 'POST':
        data = json.loads(request.body)
        product.device_type = data.get('device_type', product.device_type)
        product.model = data.get('model', product.model)
        product.basic_config = data.get('basic_config', product.basic_config)
        product.cpu = data.get('cpu', product.cpu)
        product.memory = data.get('memory', product.memory)
        product.hard_disk = data.get('hard_disk', product.hard_disk)
        product.graphic_card = data.get('graphic_card', product.graphic_card)
        product.wifi = data.get('wifi', product.wifi)
        product.configuration_parameters = data.get('configuration_parameters', product.configuration_parameters)
        product.weight = data.get('weight', product.weight)
        product.price = data.get('price', product.price)
        product.original_price = data.get('original_price', product.original_price)
        product.discount_rate = data.get('discount_rate', product.discount_rate)
        product.changing_config_price = data.get('changing_config_price', product.changing_config_price)
        product.official_link = data.get('official_link', product.official_link)
        product.image_url = data.get('image_url', product.image_url)
        product.is_active = data.get('is_active', product.is_active)
        product.save()
        return JsonResponse({'success': True, 'message': '更新成功'})
    elif request.method == 'DELETE':
        product.delete()
        return JsonResponse({'success': True, 'message': '删除成功'})

# Lenovo 产品管理
@csrf_exempt
@require_admin
def api_admin_lenovo_products(request):
    """获取/添加 Lenovo 产品列表"""
    if request.method == 'GET':
        products = LenovoProduct.objects.all().order_by('device_type', 'model', 'configuration')
        data = [{
            'id': p.id,
            'device_type': p.device_type,
            'model': p.model,
            'configuration': p.configuration,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'discount_price': str(p.discount_price),
            'original_price': str(p.original_price),
            'image_url': p.image_url,
            'is_active': p.is_active,
        } for p in products]
        return JsonResponse({'success': True, 'data': data})
    elif request.method == 'POST':
        data = json.loads(request.body)
        try:
            product = LenovoProduct.objects.create(
                device_type=data.get('device_type'),
                model=data.get('model'),
                configuration=data.get('configuration', ''),
                cpu=data.get('cpu', ''),
                memory=data.get('memory', ''),
                hard_disk=data.get('hard_disk', ''),
                graphic_card=data.get('graphic_card', ''),
                screen=data.get('screen', ''),
                wifi=data.get('wifi', ''),
                weight=data.get('weight', ''),
                accessary=data.get('accessary', ''),
                configuration_parameters=data.get('configuration_parameters', ''),
                changing_config_price=data.get('changing_config_price', ''),
                discount_price=data.get('discount_price', 0),
                original_price=data.get('original_price', 0),
                discount_rate=data.get('discount_rate', 0),
                product_link=data.get('product_link', ''),
                image_url=data.get('image_url', ''),
            )
            return JsonResponse({'success': True, 'message': '产品创建成功', 'id': product.id})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_admin
def api_admin_lenovo_product(request, product_id):
    """编辑/删除 Lenovo 产品"""
    try:
        product = LenovoProduct.objects.get(id=product_id)
    except LenovoProduct.DoesNotExist:
        return JsonResponse({'success': False, 'message': '产品不存在'})

    if request.method == 'POST':
        data = json.loads(request.body)
        product.device_type = data.get('device_type', product.device_type)
        product.model = data.get('model', product.model)
        product.configuration = data.get('configuration', product.configuration)
        product.cpu = data.get('cpu', product.cpu)
        product.memory = data.get('memory', product.memory)
        product.hard_disk = data.get('hard_disk', product.hard_disk)
        product.graphic_card = data.get('graphic_card', product.graphic_card)
        product.screen = data.get('screen', product.screen)
        product.wifi = data.get('wifi', product.wifi)
        product.weight = data.get('weight', product.weight)
        product.accessary = data.get('accessary', product.accessary)
        product.configuration_parameters = data.get('configuration_parameters', product.configuration_parameters)
        product.changing_config_price = data.get('changing_config_price', product.changing_config_price)
        product.discount_price = data.get('discount_price', product.discount_price)
        product.original_price = data.get('original_price', product.original_price)
        product.discount_rate = data.get('discount_rate', product.discount_rate)
        product.product_link = data.get('product_link', product.product_link)
        product.image_url = data.get('image_url', product.image_url)
        product.is_active = data.get('is_active', product.is_active)
        product.save()
        return JsonResponse({'success': True, 'message': '更新成功'})
    elif request.method == 'DELETE':
        product.delete()
        return JsonResponse({'success': True, 'message': '删除成功'})


# ============ 产品图片管理 API ============

@csrf_exempt
@require_admin
def api_admin_product_images(request, brand, product_type, product_id):
    """获取产品的所有图片"""
    if request.method == 'GET':
        # Fix URL routing mismatch: normalize product_type for Dell/Lenovo products
        # When URL is /api/admin/dell/product/123/images/, the generic pattern captures
        # product_type='product' instead of 'dell_product'
        if brand in ['dell', 'lenovo'] and product_type == 'product':
            product_type = f'{brand}_product'

        images = ProductImage.objects.filter(
            brand=brand,
            product_type=product_type,
            product_id=product_id
        ).order_by('image_type', 'sort_order')

        # Build gallery images array
        gallery_list = [{
            'id': img.id,
            'image_type': img.image_type,
            'url': img.image.url if img.image else '',
            'title': img.title,
            'sort_order': img.sort_order,
        } for img in images]

        # Find main image (first one with image_type='main')
        main_image = None
        for img in gallery_list:
            if img['image_type'] == 'main':
                main_image = {
                    'id': img['id'],
                    'url': img['url'],
                    'image_type': 'main',
                    'title': img['title']
                }
                break

        # If no main image found, use first gallery image as main
        if main_image is None and len(gallery_list) > 0:
            main_image = {
                'id': gallery_list[0]['id'],
                'url': gallery_list[0]['url'],
                'image_type': gallery_list[0]['image_type'],
                'title': gallery_list[0]['title']
            }

        # Return structured data matching frontend expectations
        return JsonResponse({
            'success': True,
            'data': {
                'mainImage': main_image,
                'galleryImages': gallery_list
            }
        })

    return JsonResponse({'success': False, 'message': '不支持的请求方法'})


@csrf_exempt
@require_admin
def api_admin_upload_image(request, brand):
    """上传产品图片"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支持 POST 请求'})
    
    try:
        # 获取表单数据
        product_id = request.POST.get('product_id')
        product_type = request.POST.get('product_type')
        image_type = request.POST.get('image_type', 'gallery')
        title = request.POST.get('title', '')
        sort_order = int(request.POST.get('sort_order', 0))
        
        # 获取上传的文件
        if 'image' not in request.FILES:
            return JsonResponse({'success': False, 'message': '请选择要上传的图片'})
        
        image_file = request.FILES['image']
        
        # 验证文件类型
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if image_file.content_type not in allowed_types:
            return JsonResponse({'success': False, 'message': '只支持 JPG/PNG/GIF/WebP 格式的图片'})
        
        # 如果是主图类型，确保同一产品只有一个主图
        if image_type == 'main':
            ProductImage.objects.filter(
                brand=brand,
                product_type=product_type,
                product_id=product_id,
                image_type='main'
            ).update(image_type='gallery')  # 将原有的主图降为副图
        
        # 创建图片记录
        product_image = ProductImage.objects.create(
            brand=brand,
            product_id=product_id,
            product_type=product_type,
            image_type=image_type,
            image=image_file,
            title=title,
            sort_order=sort_order,
        )
        
        return JsonResponse({
            'success': True, 
            'message': '图片上传成功',
            'data': {
                'id': product_image.id,
                'image_url': product_image.image.url,
                'image_type': product_image.image_type,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_delete_image(request, image_id):
    """删除产品图片"""
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'message': '只支持 DELETE 请求'})
    
    try:
        image = ProductImage.objects.get(id=image_id)
        
        # 删除文件
        if image.image:
            image.image.delete(save=False)
        
        # 删除记录
        image.delete()
        
        return JsonResponse({'success': True, 'message': '图片删除成功'})
    except ProductImage.DoesNotExist:
        return JsonResponse({'success': False, 'message': '图片不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_set_main_image(request, image_id):
    """设置主图"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支持 POST 请求'})
    
    try:
        image = ProductImage.objects.get(id=image_id)
        
        # 将该产品的所有图片都设为副图
        ProductImage.objects.filter(
            brand=image.brand,
            product_type=image.product_type,
            product_id=image.product_id,
            image_type='main'
        ).update(image_type='gallery')
        
        # 设置当前图片为主图
        image.image_type = 'main'
        image.save()
        
        return JsonResponse({'success': True, 'message': '主图设置成功'})
    except ProductImage.DoesNotExist:
        return JsonResponse({'success': False, 'message': '图片不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_reorder_images(request):
    """重新排序图片"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支持 POST 请求'})

    try:
        data = json.loads(request.body)
        image_orders = data.get('image_orders', [])  # 格式: [{id: 1, sort_order: 0}, {id: 2, sort_order: 1}, ...]

        for item in image_orders:
            ProductImage.objects.filter(id=item['id']).update(sort_order=item['sort_order'])

        return JsonResponse({'success': True, 'message': '排序更新成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ============ 产品图片管理 API (产品相关) ============

@csrf_exempt
@require_admin
def api_admin_product_upload_image(request, brand, product_type, product_id):
    """上传产品图片 - 针对特定产品"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支持 POST 请求'})

    try:
        # Fix URL routing mismatch: normalize product_type for Dell/Lenovo products
        if brand in ['dell', 'lenovo'] and product_type == 'product':
            product_type = f'{brand}_product'

        # 获取表单数据
        image_type = request.POST.get('image_type', 'gallery')
        title = request.POST.get('title', '')
        sort_order = int(request.POST.get('sort_order', 0))

        # 获取上传的文件
        if 'image' not in request.FILES:
            return JsonResponse({'success': False, 'message': '请选择要上传的图片'})

        image_file = request.FILES['image']

        # 验证文件类型
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if image_file.content_type not in allowed_types:
            return JsonResponse({'success': False, 'message': '只支持 JPG/PNG/GIF/WebP 格式的图片'})

        # 如果是主图类型，确保同一产品只有一个主图
        if image_type == 'main':
            ProductImage.objects.filter(
                brand=brand,
                product_type=product_type,
                product_id=product_id,
                image_type='main'
            ).update(image_type='gallery')  # 将原有的主图降为副图

        # 创建图片记录
        product_image = ProductImage.objects.create(
            brand=brand,
            product_id=product_id,
            product_type=product_type,
            image_type=image_type,
            image=image_file,
            title=title,
            sort_order=sort_order,
        )

        return JsonResponse({
            'success': True,
            'message': '图片上传成功',
            'data': {
                'id': product_image.id,
                'image_url': product_image.image.url,
                'image_type': product_image.image_type,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_product_set_main_image(request, brand, product_type, product_id, image_id):
    """设置产品的主图"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支持 POST 请求'})

    try:
        # Fix URL routing mismatch: normalize product_type for Dell/Lenovo products
        if brand in ['dell', 'lenovo'] and product_type == 'product':
            product_type = f'{brand}_product'

        # 获取图片
        image = ProductImage.objects.get(id=image_id, brand=brand, product_type=product_type, product_id=product_id)

        # 将该产品的所有图片都设为副图
        ProductImage.objects.filter(
            brand=brand,
            product_type=product_type,
            product_id=product_id,
            image_type='main'
        ).update(image_type='gallery')

        # 设置当前图片为主图
        image.image_type = 'main'
        image.save()

        return JsonResponse({'success': True, 'message': '主图设置成功'})
    except ProductImage.DoesNotExist:
        return JsonResponse({'success': False, 'message': '图片不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_product_delete_image(request, brand, product_type, product_id, image_id):
    """删除产品图片"""
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'message': '只支持 DELETE 请求'})

    try:
        image = ProductImage.objects.get(id=image_id, brand=brand, product_type=product_type, product_id=product_id)

        # 删除文件
        if image.image:
            image.image.delete(save=False)

        # 删除记录
        image.delete()

        return JsonResponse({'success': True, 'message': '图片删除成功'})
    except ProductImage.DoesNotExist:
        return JsonResponse({'success': False, 'message': '图片不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@require_admin
def api_admin_product_delete_main_image(request, brand, product_type, product_id):
    """删除产品的主图"""
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'message': '只支持 DELETE 请求'})

    try:
        # 查找主图
        main_image = ProductImage.objects.filter(
            brand=brand,
            product_type=product_type,
            product_id=product_id,
            image_type='main'
        ).first()

        if main_image:
            # 删除文件
            if main_image.image:
                main_image.image.delete(save=False)
            # 删除记录
            main_image.delete()
            return JsonResponse({'success': True, 'message': '主图删除成功'})
        else:
            return JsonResponse({'success': False, 'message': '该产品没有主图'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== Apple 产品图片管理 API ====================

# Apple Model 图片管理
@csrf_exempt
@require_admin
@require_http_methods(["GET"])
def api_admin_apple_model_images(request, model_id):
    """获取 Apple 型号的图片"""
    return api_admin_product_images(request, brand='apple', product_type='model', product_id=model_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_apple_model_upload_image(request, model_id):
    """上传 Apple 型号图片"""
    return api_admin_product_upload_image(request, brand='apple', product_type='model', product_id=model_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_apple_model_set_main_image(request, model_id):
    """设置 Apple 型号的主图"""
    data = json.loads(request.body)
    image_id = data.get('image_id')
    if not image_id:
        return JsonResponse({'success': False, 'message': '请提供图片ID'})
    return api_admin_product_set_main_image(request, brand='apple', product_type='model', product_id=model_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_apple_model_delete_image(request, model_id, image_id):
    """删除 Apple 型号的图片"""
    return api_admin_product_delete_image(request, brand='apple', product_type='model', product_id=model_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_apple_model_delete_main_image(request, model_id):
    """删除 Apple 型号的主图"""
    return api_admin_product_delete_main_image(request, brand='apple', product_type='model', product_id=model_id)


# Apple Config 图片管理
@csrf_exempt
@require_admin
@require_http_methods(["GET"])
def api_admin_apple_config_images(request, config_id):
    """获取 Apple 配置的图片"""
    return api_admin_product_images(request, brand='apple', product_type='config', product_id=config_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_apple_config_upload_image(request, config_id):
    """上传 Apple 配置图片"""
    return api_admin_product_upload_image(request, brand='apple', product_type='config', product_id=config_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_apple_config_set_main_image(request, config_id):
    """设置 Apple 配置的主图"""
    data = json.loads(request.body)
    image_id = data.get('image_id')
    if not image_id:
        return JsonResponse({'success': False, 'message': '请提供图片ID'})
    return api_admin_product_set_main_image(request, brand='apple', product_type='config', product_id=config_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_apple_config_delete_image(request, config_id, image_id):
    """删除 Apple 配置的图片"""
    return api_admin_product_delete_image(request, brand='apple', product_type='config', product_id=config_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_apple_config_delete_main_image(request, config_id):
    """删除 Apple 配置的主图"""
    return api_admin_product_delete_main_image(request, brand='apple', product_type='config', product_id=config_id)


# ==================== Dell 产品图片管理 API ====================

@csrf_exempt
@require_admin
@require_http_methods(["GET"])
def api_admin_dell_product_images(request, product_id):
    """获取 Dell 产品的图片"""
    return api_admin_product_images(request, brand='dell', product_type='dell_product', product_id=product_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_dell_product_upload_image(request, product_id):
    """上传 Dell 产品图片"""
    return api_admin_product_upload_image(request, brand='dell', product_type='dell_product', product_id=product_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_dell_product_set_main_image(request, product_id):
    """设置 Dell 产品的主图"""
    data = json.loads(request.body)
    image_id = data.get('image_id')
    if not image_id:
        return JsonResponse({'success': False, 'message': '请提供图片ID'})
    return api_admin_product_set_main_image(request, brand='dell', product_type='dell_product', product_id=product_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_dell_product_delete_image(request, product_id, image_id):
    """删除 Dell 产品的图片"""
    return api_admin_product_delete_image(request, brand='dell', product_type='dell_product', product_id=product_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_dell_product_delete_main_image(request, product_id):
    """删除 Dell 产品的主图"""
    return api_admin_product_delete_main_image(request, brand='dell', product_type='dell_product', product_id=product_id)


# ==================== Lenovo 产品图片管理 API ====================

@csrf_exempt
@require_admin
@require_http_methods(["GET"])
def api_admin_lenovo_product_images(request, product_id):
    """获取 Lenovo 产品的图片"""
    return api_admin_product_images(request, brand='lenovo', product_type='lenovo_product', product_id=product_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_lenovo_product_upload_image(request, product_id):
    """上传 Lenovo 产品图片"""
    return api_admin_product_upload_image(request, brand='lenovo', product_type='lenovo_product', product_id=product_id)


@csrf_exempt
@require_admin
@require_http_methods(["POST"])
def api_admin_lenovo_product_set_main_image(request, product_id):
    """设置 Lenovo 产品的主图"""
    data = json.loads(request.body)
    image_id = data.get('image_id')
    if not image_id:
        return JsonResponse({'success': False, 'message': '请提供图片ID'})
    return api_admin_product_set_main_image(request, brand='lenovo', product_type='lenovo_product', product_id=product_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_lenovo_product_delete_image(request, product_id, image_id):
    """删除 Lenovo 产品的图片"""
    return api_admin_product_delete_image(request, brand='lenovo', product_type='lenovo_product', product_id=product_id, image_id=image_id)


@csrf_exempt
@require_admin
@require_http_methods(["DELETE"])
def api_admin_lenovo_product_delete_main_image(request, product_id):
    """删除 Lenovo 产品的主图"""
    return api_admin_product_delete_main_image(request, brand='lenovo', product_type='lenovo_product', product_id=product_id)


# ==================== Dell 选配中心 ====================


def dell_index(request):
    """Dell选配中心首页"""
    return render(request, 'dell/index.html')


def dell_notebook(request):
    """Dell笔记本选配"""
    products = DellProduct.objects.filter(device_type='Notebook', is_active=True)
    user_name = request.session.get('user_name', '')
    
    import json
    products_json = json.dumps([
        {
            'model': p.model,
            'basic_config': p.basic_config,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'wifi': p.wifi,
            'configuration_parameters': p.configuration_parameters,
            'weight': p.weight,
            'price': float(p.price),
            'original_price': float(p.original_price),
            'discount_rate': float(p.discount_rate) if p.discount_rate else 0,
            'changing_config_price': p.changing_config_price,
            'official_link': p.official_link,
            'remark': p.remark,
        }
        for p in products
    ], ensure_ascii=False)
    
    return render(request, 'dell/notebook.html', {
        'products': products,
        'products_json': products_json,
        'brand': 'dell',
        'user_name': user_name,
    })


def dell_laptop(request):
    """Dell高性能本选配"""
    products = DellProduct.objects.filter(device_type='Laptop', is_active=True)
    user_name = request.session.get('user_name', '')
    
    import json
    products_json = json.dumps([
        {
            'model': p.model,
            'basic_config': p.basic_config,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'wifi': p.wifi,
            'configuration_parameters': p.configuration_parameters,
            'weight': p.weight,
            'price': float(p.price),
            'original_price': float(p.original_price),
            'discount_rate': float(p.discount_rate) if p.discount_rate else 0,
            'changing_config_price': p.changing_config_price,
            'official_link': p.official_link,
            'remark': p.remark,
        }
        for p in products
    ], ensure_ascii=False)
    
    return render(request, 'dell/laptop.html', {
        'products': products,
        'products_json': products_json,
        'brand': 'dell',
        'user_name': user_name,
    })


def dell_desktop(request):
    """Dell台式机选配"""
    products = DellProduct.objects.filter(device_type='Desktop', is_active=True)
    user_name = request.session.get('user_name', '')
    
    import json
    products_json = json.dumps([
        {
            'model': p.model,
            'basic_config': p.basic_config,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'wifi': p.wifi,
            'configuration_parameters': p.configuration_parameters,
            'weight': p.weight,
            'price': float(p.price),
            'original_price': float(p.original_price),
            'discount_rate': float(p.discount_rate) if p.discount_rate else 0,
            'changing_config_price': p.changing_config_price,
            'official_link': p.official_link,
            'remark': p.remark,
        }
        for p in products
    ], ensure_ascii=False)
    
    return render(request, 'dell/desktop.html', {
        'products': products,
        'products_json': products_json,
        'brand': 'dell',
        'user_name': user_name,
    })


def dell_monitor(request):
    """Dell显示器选配"""
    products = DellProduct.objects.filter(device_type='Monitor', is_active=True)
    user_name = request.session.get('user_name', '')
    
    import json
    products_json = json.dumps([
        {
            'model': p.model,
            'basic_config': p.basic_config,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'wifi': p.wifi,
            'configuration_parameters': p.configuration_parameters,
            'weight': p.weight,
            'price': float(p.price),
            'original_price': float(p.original_price),
            'discount_rate': float(p.discount_rate) if p.discount_rate else 0,
            'changing_config_price': p.changing_config_price,
            'official_link': p.official_link,
            'remark': p.remark,
        }
        for p in products
    ], ensure_ascii=False)
    
    return render(request, 'dell/monitor.html', {
        'products': products,
        'products_json': products_json,
        'brand': 'dell',
        'user_name': user_name,
    })


def dell_accessories(request):
    """Dell配件选配"""
    products = DellProduct.objects.filter(device_type='Docking', is_active=True)
    user_name = request.session.get('user_name', '')
    
    import json
    products_json = json.dumps([
        {
            'model': p.model,
            'basic_config': p.basic_config,
            'cpu': p.cpu,
            'memory': p.memory,
            'hard_disk': p.hard_disk,
            'graphic_card': p.graphic_card,
            'wifi': p.wifi,
            'configuration_parameters': p.configuration_parameters,
            'weight': p.weight,
            'price': float(p.price),
            'original_price': float(p.original_price),
            'discount_rate': float(p.discount_rate) if p.discount_rate else 0,
            'changing_config_price': p.changing_config_price,
            'official_link': p.official_link,
            'remark': p.remark,
        }
        for p in products
    ], ensure_ascii=False)
    
    return render(request, 'dell/accessories.html', {
        'products': products,
        'products_json': products_json,
        'brand': 'dell',
        'user_name': user_name,
    })


def dell_detail(request, model_name):
    """Dell产品详情页"""
    try:
        product = DellProduct.objects.get(model=model_name, is_active=True)
    except DellProduct.DoesNotExist:
        return render(request, '404.html', status=404)
    
    # 计算折扣百分比
    original_price = float(product.original_price)
    price = float(product.price)
    if original_price and original_price > price:
        discount_percent = int((1 - price / original_price) * 100)
    else:
        discount_percent = 0
    
    # 获取产品图片 - 从 ProductImage 模型查询
    main_image = None
    thumbnails = []
    
    # 查询该产品的图片
    images = ProductImage.objects.filter(
        brand='dell',
        product_type='dell_product',
        product_id=product.id
    ).order_by('sort_order', 'created_at')
    
    if images.exists():
        # 获取主图（优先显示主图，然后是第一张图）
        main_img_obj = images.filter(image_type='main').first() or images.first()
        if main_img_obj and main_img_obj.image:
            main_image = main_img_obj.image.url
        
        # 获取所有图片作为缩略图
        for img in images:
            if img.image:
                thumbnails.append(img.image.url)
    else:
        # 如果没有上传图片，使用产品的 image_url 作为后备
        if product.image_url:
            main_image = product.image_url
    
    import json
    product_data = {
        'model': product.model,
        'basic_config': product.basic_config,
        'device_type': product.device_type,
        'cpu': product.cpu,
        'memory': product.memory,
        'hard_disk': product.hard_disk,
        'graphic_card': product.graphic_card,
        'wifi': product.wifi,
        'configuration_parameters': product.configuration_parameters,
        'weight': product.weight,
        'price': price,
        'original_price': original_price,
        'discount_rate': float(product.discount_rate) if product.discount_rate else 0,
        'changing_config_price': product.changing_config_price,
        'official_link': product.official_link,
        'remark': product.remark,
    }
    product_json = json.dumps(product_data, ensure_ascii=False)
    user_name = request.session.get('user_name', '')
    
    # 处理configuration_parameters - 转换为HTML格式
    configuration_parameters_html = ''
    if product.configuration_parameters:
        params_text = product.configuration_parameters
        
        # 解析端口信息（格式如 "right:\n1.xxx\n2.xxx\n\nleft:\n1.xxx"）
        sections = params_text.split('\n\n')
        configuration_parameters_html = '<div class="port-info">'
        
        for section in sections:
            section = section.strip()
            if not section:
                continue
            
            # 解析section标题和内容
            lines = section.split('\n')
            if len(lines) >= 2:
                title = lines[0].strip(':').strip()
                items = lines[1:]
                
                configuration_parameters_html += f'<div class="port-section"><h5>{title}</h5>'
                configuration_parameters_html += '<ul class="port-list">'
                for item in items:
                    item = item.strip()
                    if item and (item[0].isdigit() or item.startswith('-')):
                        # 移除开头的数字序号
                        item = item.lstrip('0123456789. ')
                        configuration_parameters_html += f'<li>{item}</li>'
                configuration_parameters_html += '</ul></div>'
            elif len(lines) == 1:
                # 直接显示文本
                configuration_parameters_html += f'<p>{lines[0]}</p>'
        
        configuration_parameters_html += '</div>'
    
    # 处理changing_config_price - 转换为HTML格式
    configuration_upgrade_html = ''
    if product.changing_config_price:
        # 按行分割，处理不同的配置选项
        lines = product.changing_config_price.split('\n')
        configuration_upgrade_html = '<div class="upgrade-list">'
        
        for line in lines:
            line = line.strip()
            if line:
                # 处理特殊符号开头的内容
                upgrade_name = line
                if line.startswith('★'):
                    upgrade_name = line[1:].strip()
                
                # 尝试提取价格（格式如 "+1200 yuan" 或 "，+1200 yuan"）
                price_parts = upgrade_name.rsplit('，+', 1)
                if len(price_parts) == 2:
                    name = price_parts[0].strip()
                    price = '+' + price_parts[1].strip()
                    configuration_upgrade_html += f'''<div class="upgrade-item">
                        <div class="upgrade-checkbox"></div>
                        <div class="upgrade-info">
                            <span class="upgrade-name">{name}</span>
                            <span class="upgrade-price">{price}</span>
                        </div>
                    </div>'''
                elif '+' in upgrade_name:
                    # 直接在末尾的格式
                    parts = upgrade_name.rsplit('+', 1)
                    if len(parts) == 2:
                        name = parts[0].strip()
                        price = '+' + parts[1].strip()
                        configuration_upgrade_html += f'''<div class="upgrade-item">
                            <div class="upgrade-checkbox"></div>
                            <div class="upgrade-info">
                                <span class="upgrade-name">{name}</span>
                                <span class="upgrade-price">{price}</span>
                            </div>
                        </div>'''
                    else:
                        configuration_upgrade_html += f'<div class="upgrade-item">{upgrade_name}</div>'
                else:
                    configuration_upgrade_html += f'<div class="upgrade-item">{upgrade_name}</div>'
        
        configuration_upgrade_html += '</div>'
    
    return render(request, 'dell/detail.html', {
        'product': product_data,
        'product_json': product_json,
        'device_type': product.device_type,
        'brand': 'dell',
        'user_name': user_name,
        'discount_percent': discount_percent,
        'configuration_parameters_html': configuration_parameters_html,
        'configuration_upgrade_html': configuration_upgrade_html,
        'main_image': main_image,
        'thumbnails': thumbnails,
    })


# ==================== Lenovo 选配中心 ====================


def lenovo_index(request):
    """Lenovo选配中心首页"""
    return render(request, 'lenovo/index.html')


def lenovo_laptop(request):
    """Lenovo笔记本选配"""
    products = LenovoProduct.objects.filter(device_type='Laptop', is_active=True)
    user_name = request.session.get('user_name', '')
    
    # 获取所有型号用于侧边栏
    models = products.values('model').distinct()
    
    # 获取选中的型号
    selected_model = request.GET.get('model', '')
    configurations = []
    selected_product = None
    
    if selected_model:
        configurations = list(products.filter(model=selected_model))
        if configurations:
            selected_product = configurations[0]
    
    # 如果没有选中型号，默认显示第一个型号的配置
    if not selected_model and products.exists():
        selected_model = products.first().model
        configurations = list(products.filter(model=selected_model))
        selected_product = configurations[0] if configurations else None
    
    return render(request, 'lenovo/laptop.html', {
        'products': products,
        'models': models,
        'selected_model': {'model': selected_model} if selected_model else None,
        'selected_product': selected_product,
        'configurations': configurations,
        'brand': 'lenovo',
        'product_type': 'Laptop',
        'user_name': user_name,
    })


def lenovo_desktop(request):
    """Lenovo台式机选配"""
    products = LenovoProduct.objects.filter(device_type='Desktop', is_active=True)
    user_name = request.session.get('user_name', '')
    
    # 获取所有型号用于侧边栏
    models = products.values('model').distinct()
    
    # 获取选中的型号
    selected_model = request.GET.get('model', '')
    configurations = []
    selected_product = None
    
    if selected_model:
        configurations = list(products.filter(model=selected_model))
        if configurations:
            selected_product = configurations[0]
    
    # 如果没有选中型号，默认显示第一个型号的配置
    if not selected_model and products.exists():
        selected_model = products.first().model
        configurations = list(products.filter(model=selected_model))
        selected_product = configurations[0] if configurations else None
    
    return render(request, 'lenovo/desktop.html', {
        'products': products,
        'models': models,
        'selected_model': {'model': selected_model} if selected_model else None,
        'selected_product': selected_product,
        'configurations': configurations,
        'brand': 'lenovo',
        'product_type': 'Desktop',
        'user_name': user_name,
    })


def lenovo_detail(request, config_id):
    """Lenovo产品详情页"""
    try:
        config = LenovoProduct.objects.get(id=config_id, is_active=True)
    except LenovoProduct.DoesNotExist:
        return render(request, '404.html', status=404)
    
    # 计算折扣百分比
    original_price = float(config.original_price) if config.original_price else 0
    price = float(config.discount_price) if config.discount_price else 0
    if original_price and original_price > price:
        discount_percent = int((1 - price / original_price) * 100)
    else:
        discount_percent = 0
    
    # 获取产品图片 - 从 ProductImage 模型查询
    main_image = None
    thumbnails = []
    
    # 查询该产品的图片
    images = ProductImage.objects.filter(
        brand='lenovo',
        product_type='lenovo_product',
        product_id=config.id
    ).order_by('sort_order', 'created_at')
    
    if images.exists():
        # 获取主图（优先显示主图，然后是第一张图）
        main_img_obj = images.filter(image_type='main').first() or images.first()
        if main_img_obj and main_img_obj.image:
            main_image = main_img_obj.image.url
        
        # 获取所有图片作为缩略图
        for img in images:
            if img.image:
                thumbnails.append(img.image.url)
    else:
        # 如果没有上传图片，使用产品的 image_url 作为后备
        if config.image_url:
            main_image = config.image_url
    
    user_name = request.session.get('user_name', '')
    
    # 处理configuration_parameters - 转换为HTML格式
    configuration_parameters_html = ''
    if config.configuration_parameters:
        params_text = config.configuration_parameters
        
        # 尝试解析端口信息
        sections = params_text.split('\n\n')
        configuration_parameters_html = '<div class="port-info">'
        
        for section in sections:
            section = section.strip()
            if not section:
                continue
            
            lines = section.split('\n')
            if len(lines) >= 2:
                title = lines[0].strip(':').strip()
                items = lines[1:]
                
                configuration_parameters_html += f'<div class="port-section"><h5>{title}</h5>'
                configuration_parameters_html += '<ul class="port-list">'
                for item in items:
                    item = item.strip()
                    if item and (item[0].isdigit() or item.startswith('-')):
                        item = item.lstrip('0123456789. ')
                        configuration_parameters_html += f'<li>{item}</li>'
                configuration_parameters_html += '</ul></div>'
            elif len(lines) == 1:
                configuration_parameters_html += f'<p>{lines[0]}</p>'
        
        configuration_parameters_html += '</div>'
    
    # 处理changing_config_price - 转换为HTML格式
    configuration_upgrade_html = ''
    if config.changing_config_price:
        lines = config.changing_config_price.split('\n')
        configuration_upgrade_html = '<div class="upgrade-list">'
        
        for line in lines:
            line = line.strip()
            if line:
                upgrade_name = line
                if line.startswith('★'):
                    upgrade_name = line[1:].strip()
                
                price_parts = upgrade_name.rsplit('，+', 1)
                if len(price_parts) == 2:
                    name = price_parts[0].strip()
                    price_str = '+' + price_parts[1].strip()
                    configuration_upgrade_html += f'''<div class="upgrade-item">
                        <div class="upgrade-checkbox"></div>
                        <div class="upgrade-info">
                            <span class="upgrade-name">{name}</span>
                            <span class="upgrade-price">{price_str}</span>
                        </div>
                    </div>'''
                elif '+' in upgrade_name:
                    parts = upgrade_name.rsplit('+', 1)
                    if len(parts) == 2:
                        name = parts[0].strip()
                        price_str = '+' + parts[1].strip()
                        configuration_upgrade_html += f'''<div class="upgrade-item">
                            <div class="upgrade-checkbox"></div>
                            <div class="upgrade-info">
                                <span class="upgrade-name">{name}</span>
                                <span class="upgrade-price">{price_str}</span>
                            </div>
                        </div>'''
                    else:
                        configuration_upgrade_html += f'<div class="upgrade-item">{upgrade_name}</div>'
                else:
                    configuration_upgrade_html += f'<div class="upgrade-item">{upgrade_name}</div>'
        
        configuration_upgrade_html += '</div>'
    
    # 获取设备类型名称
    device_type = config.device_type
    device_type_name = '笔记本' if device_type == 'Laptop' else '台式机'
    
    return render(request, 'lenovo/detail.html', {
        'configuration': config,
        'device_type': device_type,
        'device_type_name': device_type_name,
        'brand': 'lenovo',
        'user_name': user_name,
        'discount_percent': discount_percent,
        'configuration_parameters_html': configuration_parameters_html,
        'configuration_upgrade_html': configuration_upgrade_html,
        'main_image': main_image,
        'thumbnails': thumbnails,
    })